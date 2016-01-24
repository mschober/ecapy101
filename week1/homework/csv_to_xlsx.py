import csv
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
import sys

src_file = sys.argv[1]
tgt_file = sys.argv[2]

dest_filename = r"{0}".format(tgt_file)

f = open(r'{0}'.format(src_file), 'rU')

#csv.register_dialect('colons', delimiter=':')

#reader = csv.reader(f, dialect='colons')
reader = csv.reader(f)

wb = Workbook()

ws = wb.worksheets[0]
ws.title = "output.xlsx"

print "The source file is %s" % src_file
print "The target file is %s" % tgt_file

print "Begin converting to xlsx..."
for row_index, row in enumerate(reader):
    for column_index, cell in enumerate(row):
        column_letter = get_column_letter((column_index + 1))
        ws.cell('%s%s'%(column_letter, (row_index + 1))).value = cell

print "Saving the converted file..."
wb.save(filename = dest_filename)

print "Success!"
